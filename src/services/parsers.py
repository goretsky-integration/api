import pathlib
import unicodedata
import uuid
import operator
from abc import ABC, abstractmethod
from typing import Any

import openpyxl
from bs4 import BeautifulSoup
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from models.external_api_responses.office_manager.accounting import StockBalance
from models.external_api_responses.office_manager import (
    StopSaleBySector,
    StopSaleByStreet,
    TripsWithOneOrder,
    UnitDeliveryPartialStatistics,
    UnitKitchenPartialStatistics,
)
from models.external_api_responses.shift_manager import OrderPartial, OrderByUUID
from models.external_api_responses.export_service_api import UsedPromoCode

__all__ = (
    'PartialStatisticsParser',
    'HTMLParser',
    'SectorStopSalesHTMLParser',
    'StreetStopSalesHTMLParser',
    'DeliveryStatisticsHTMLParser',
    'KitchenStatisticsHTMLParser',
    'StockBalanceHTMLParser',
    'OrderByUUIDParser',
    'OrdersPartial',
    'ExcelParser',
    'DeliveryStatisticsExcelParser',
    'UsedPromoCodesExcelParser',
)


class HTMLParser(ABC):

    def __init__(self, html: str):
        self._html = html
        self._soup = BeautifulSoup(html, 'lxml')

    @abstractmethod
    def parse(self) -> Any:
        pass

    @staticmethod
    def clear_extra_symbols(text: str) -> str:
        text = unicodedata.normalize('NFKD', text)
        for i in (' ', '₽', '%', '\r', '\t'):
            text = text.replace(i, '')
        return text.strip().replace(',', '.').replace('−', '-')


class PartialStatisticsParser(HTMLParser):

    def __init__(self, html: str, unit_id: int | str):
        super().__init__(html)
        self._unit_id = unit_id
        self._panel_titles = self.parse_panel_titles()

    @abstractmethod
    def parse_panel_titles(self) -> list[str]:
        pass


class SectorStopSalesHTMLParser(HTMLParser):
    def parse(self) -> list[StopSaleBySector]:
        trs = self._soup.find('table', id='bootgrid-table').find('tbody').find_all('tr')
        nested_trs = [[td.text.strip() for td in tr.find_all('td')] for tr in trs]
        return [
            StopSaleBySector(
                unit_name=tds[0],
                sector=tds[1],
                started_at=tds[2],
                staff_name_who_stopped=tds[3],
                staff_name_who_resumed=tds[5],
            ) for tds in nested_trs
        ]


class StreetStopSalesHTMLParser(HTMLParser):
    def parse(self) -> list[StopSaleByStreet]:
        trs = self._soup.find('table', id='bootgrid-table').find_all('tr')[1:]
        nested_trs = [[td.text.strip() for td in tr.find_all('td')] for tr in trs]
        return [
            StopSaleByStreet(
                unit_name=tds[0],
                started_at=tds[3],
                staff_name_who_stopped=tds[4],
                staff_name_who_resumed=tds[6],
                sector=tds[1],
                street=tds[2],
            ) for tds in nested_trs
        ]


class DeliveryStatisticsHTMLParser(PartialStatisticsParser):

    def parse_panel_titles(self) -> list[str]:
        return [self.clear_extra_symbols(i.text)
                for i in self._soup.find_all('h1', class_='operationalStatistics_panelTitle')]

    def parse(self) -> UnitDeliveryPartialStatistics:
        couriers_on_shift_count, couriers_in_queue_count = self._panel_titles[3].split('/')
        return UnitDeliveryPartialStatistics(
            unit_id=self._unit_id,
            heated_shelf_orders_count=self._panel_titles[2],
            couriers_in_queue_count=couriers_in_queue_count,
            couriers_on_shift_count=couriers_on_shift_count,
        )


class KitchenStatisticsHTMLParser(PartialStatisticsParser):
    __slots__ = ('_soup',)

    def parse_panel_titles(self) -> list[str]:
        return [self.clear_extra_symbols(i.text)
                for i in self._soup.find_all('h1', class_='operationalStatistics_panelTitle')]

    def parse(self) -> UnitKitchenPartialStatistics:
        sales_per_labor_hour_today, from_week_before_percent = self._panel_titles[0].split('\n')
        minutes, seconds = map(int, self._panel_titles[3].split(':'))
        total_cooking_time = minutes * 60 + seconds
        return UnitKitchenPartialStatistics(
            unit_id=self._unit_id,
            sales_per_labor_hour_today=sales_per_labor_hour_today,
            from_week_before_percent=from_week_before_percent,
            total_cooking_time=total_cooking_time,
        )


class StockBalanceHTMLParser(HTMLParser):

    def __init__(self, html: str, unit_id: int):
        super().__init__(html)
        self.unit_id = unit_id

    def parse(self) -> list[StockBalance]:
        trs = self._soup.find('tbody').find_all('tr')
        result: list[StockBalance] = []
        for tr in trs:
            tds = tr.find_all('td')
            if len(tds) != 6:
                continue
            ingredient_name, stocks_count, _, _, _, days_left = [td.text.strip() for td in tds]
            if not days_left.isdigit():
                continue
            *ingredient_name_parts, stocks_unit = ingredient_name.split(',')
            ingredient_name = ','.join(ingredient_name_parts)
            result.append(StockBalance(
                unit_id=self.unit_id,
                ingredient_name=ingredient_name,
                days_left=days_left,
                stocks_unit=stocks_unit.strip(),
                stocks_count=stocks_count.strip().replace(',', '.').replace(' ', ''),
            ))
        return result


class OrdersPartial(HTMLParser):

    def parse(self) -> list[OrderPartial]:
        trs = self._soup.find_all('tr')[1:]
        nested_trs = [tr.find_all('td') for tr in trs]
        return [
            OrderPartial(
                uuid=td[0].find('a').get('href').split('=')[-1],
                number=td[1].text.strip(),
                price=td[4].text.strip('₽').strip(),
                type=td[7].text
            ) for td in nested_trs
        ]


class OrderByUUIDParser(HTMLParser):

    def __init__(self, html: str, order_uuid: uuid.UUID, order_price: int, order_type: str):
        super().__init__(html)
        self._order_uuid = order_uuid
        self._order_price = order_price
        self._order_type = order_type

    def parse(self) -> OrderByUUID:
        order_no = self._soup.find('span', id='orderNumber').text
        department = self._soup.find('div', class_='headerDepartment').text

        courier_name: str | None = None
        for tr in self._soup.find('table').find_all('tr'):
            if len(tr) != 2:
                continue
            field_name, field_value = [td.text.strip() for td in tr.find_all('td')]
            if field_name == 'Курьер:' and field_value:
                courier_name = field_value
                break

        history = self._soup.find('div', id='history')
        trs = history.find_all('tr')[1:]
        order_created_at = receipt_printed_at = order_canceled_at = None
        is_receipt_printed = False
        for tr in trs:
            _, msg, _ = tr.find_all('td')
            msg = msg.text.lower().strip()
            if 'закрыт чек на возврат' in msg:
                is_receipt_printed = True
                break

        rejected_by_user_name: str | None = None
        for tr in trs:
            dt, msg, user_name = tr.find_all('td')
            msg = msg.text.lower().strip()
            if 'has been accepted' in msg:
                order_created_at = dt.text
            elif 'закрыт чек на возврат' in msg and is_receipt_printed:
                receipt_printed_at = dt.text
            elif 'has been rejected' in msg:
                order_canceled_at = dt.text
                if user_name.text.strip():
                    rejected_by_user_name = user_name.text.strip()

        return OrderByUUID(
            number=order_no,
            unit_name=department,
            created_at=order_created_at,
            canceled_at=order_canceled_at,
            receipt_printed_at=receipt_printed_at,
            uuid=self._order_uuid,
            price=self._order_price,
            type=self._order_type,
            courier_name=courier_name,
            rejected_by_user_name=rejected_by_user_name,
        )


class ExcelParser(ABC):

    def __init__(self, file_path: str | pathlib.Path):
        self._wb: Workbook = openpyxl.load_workbook(file_path, read_only=True)
        self._ws: Worksheet = self._wb.active

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()

    def close(self) -> None:
        self._wb.close()

    @abstractmethod
    def parse(self):
        pass


class DeliveryStatisticsExcelParser(ExcelParser):

    def parse(self) -> list[TripsWithOneOrder]:
        rows = self._ws['A7': f'N{self._ws.max_row}']
        result = []
        for row in rows:
            unit_name, *_, trips_with_one_order_percentage = [cell.value for cell in row]
            result.append(TripsWithOneOrder(
                unit_name=unit_name,
                percentage=round(trips_with_one_order_percentage * 100, 2))
            )
        return result


class UsedPromoCodesExcelParser(ExcelParser):
    def parse(self) -> list[UsedPromoCode]:
        rows = self._ws['A6': f'I{self._ws.max_row}']
        used_promo_codes: list[UsedPromoCode] = []
        for row in rows:
            values = [value for cell in row if (value := cell.value) is not None]
            if len(values) != 9:
                continue
            (
                unit_name,
                promo_code,
                event,
                typical_description,
                order_type,
                order_status,
                order_no,
                ordered_at,
                order_price,
            ) = values

            used_promo_codes.append(
                UsedPromoCode(
                    unit_name=unit_name,
                    promo_code=promo_code,
                    event=event,
                    typical_description=typical_description,
                    order_type=order_type,
                    order_status=order_status,
                    order_no=order_no,
                    ordered_at=ordered_at,
                    order_price=order_price,
                )
            )
        return used_promo_codes
